"""Génération du classeur Excel récapitulatif, mise en forme professionnelle."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excel_config import COLUMNS
from state import InvoiceRow

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

TITLE_ROW = 1
HEADER_ROW = 3
DATA_START_ROW = 4


def _apply_number_format(cell, fmt: str) -> None:
    if fmt == "amount":
        cell.number_format = "#,##0"
    elif fmt == "percent_int":
        cell.number_format = "0"
    elif fmt == "date":
        cell.number_format = "DD/MM/YYYY"


def build_workbook(rows: list[InvoiceRow], month_label: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan mensuel"

    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=len(COLUMNS))
    title_cell = ws.cell(
        row=TITLE_ROW, column=1,
        value=f"O2 Coco — Récapitulatif des factures fournisseurs — {month_label}",
    )
    title_cell.font = Font(size=14, bold=True)

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col["header"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row in enumerate(rows):
        r = DATA_START_ROW + row_offset
        for col_idx, col in enumerate(COLUMNS, start=1):
            value = getattr(row, col["field"])
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER
            _apply_number_format(cell, col["format"])

    total_row = DATA_START_ROW + len(rows)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = BORDER
        cell.font = Font(bold=True)
        if col.get("totals"):
            cell.value = sum(getattr(row, col["field"]) for row in rows)
            cell.number_format = "#,##0"

    for col_idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_filename(month_key: str) -> str:
    return f"Bilan_O2Coco_{month_key}.xlsx"
